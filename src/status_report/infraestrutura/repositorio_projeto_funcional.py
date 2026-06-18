"""Leitura da planilha 'Projeto Funcional' de um cliente.

A planilha de origem pode estar em dois formatos no Drive:
- Google Sheets nativo  -> lido via Sheets API;
- arquivo Excel (.xlsx)  -> baixado via Drive API e lido com openpyxl.

O codigo detecta o tipo automaticamente, entao a coordenacao pode colar o link
sem se preocupar com o formato.

O mapeamento das colunas e feito pelo NOME do cabecalho (e nao pela posicao),
porque a planilha pode conter colunas ocultas ou extras (Nivel, Data Entrega,
Avaliacao, etc.) que variam entre clientes.
"""
from __future__ import annotations

import datetime
import io
import re
import unicodedata

import openpyxl

from status_report.dominio.modelos import ItemProjetoFuncional
from status_report.infraestrutura.repositorio_planilha import ler_intervalo

_RE_ID_PLANILHA = re.compile(r"/spreadsheets/d/([a-zA-Z0-9_-]+)")
_RE_ID_SOLTO = re.compile(r"^[a-zA-Z0-9_-]{20,}$")

# Quantas linhas/colunas varremos ao ler a planilha de origem.
_MAX_LINHAS = 300
_RANGE_LEITURA = "A1:Z300"

_MIME_GOOGLE_SHEET = "application/vnd.google-apps.spreadsheet"
_MIMES_EXCEL = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
}


class PlanilhaProjetoInvalida(RuntimeError):
    """Erro ao localizar/ler a planilha de Projeto Funcional."""


def extrair_id_planilha(link_ou_id: str) -> str:
    """Aceita um link completo do Sheets ou um ID puro e retorna o ID."""
    valor = (link_ou_id or "").strip()
    if not valor:
        raise PlanilhaProjetoInvalida("Link da planilha de Projeto Funcional vazio.")

    correspondencia = _RE_ID_PLANILHA.search(valor)
    if correspondencia:
        return correspondencia.group(1)

    if _RE_ID_SOLTO.match(valor):
        return valor

    raise PlanilhaProjetoInvalida(
        f"Nao consegui extrair o ID da planilha a partir de: {valor!r}. "
        "Cole o link completo do Google Sheets ou apenas o ID."
    )


def ler_itens_projeto_funcional(
    sheets,
    drive,
    spreadsheet_id: str,
    aba: str | None = None,
) -> tuple[str, list[ItemProjetoFuncional]]:
    """Le e converte a planilha em itens tipados, qualquer que seja o formato.

    Returns:
        (nome_da_aba, lista_de_itens)
    """
    mimetype, nome_arquivo = _obter_metadados(drive, spreadsheet_id)

    if mimetype == _MIME_GOOGLE_SHEET:
        aba, linhas = _ler_linhas_google_sheet(sheets, spreadsheet_id, aba)
    elif mimetype in _MIMES_EXCEL:
        aba, linhas = _ler_linhas_excel(drive, spreadsheet_id, aba)
    else:
        raise PlanilhaProjetoInvalida(
            f"Tipo de arquivo nao suportado para '{nome_arquivo}' ({mimetype}). "
            "Use um Google Sheets ou um arquivo Excel (.xlsx)."
        )

    idx_cabecalho = _localizar_linha_cabecalho(linhas)
    if idx_cabecalho is None:
        raise PlanilhaProjetoInvalida(
            f"Nao encontrei a linha de cabecalho (com 'Status' e 'Descricao') "
            f"na aba '{aba}'."
        )

    mapa = _mapear_colunas(linhas[idx_cabecalho])

    itens: list[ItemProjetoFuncional] = []
    for linha in linhas[idx_cabecalho + 1 :]:
        item = _linha_para_item(linha, mapa, sequencial=len(itens) + 1)
        if item is not None:
            itens.append(item)
    return aba, itens


# ---------------------------------------------------------------------------
# Leitura de baixo nivel (por formato)
# ---------------------------------------------------------------------------


def _obter_metadados(drive, spreadsheet_id: str) -> tuple[str, str]:
    meta = (
        drive.files()
        .get(fileId=spreadsheet_id, fields="mimeType,name", supportsAllDrives=True)
        .execute()
    )
    return meta.get("mimeType", ""), meta.get("name", "")


def _ler_linhas_google_sheet(
    sheets, spreadsheet_id: str, aba: str | None
) -> tuple[str, list[list[str]]]:
    def carregar(titulo: str) -> list[list[str]]:
        return ler_intervalo(
            sheets=sheets,
            spreadsheet_id=spreadsheet_id,
            intervalo=f"'{titulo}'!{_RANGE_LEITURA}",
        )

    if aba is not None:
        return aba, carregar(aba)

    meta = (
        sheets.spreadsheets()
        .get(spreadsheetId=spreadsheet_id, fields="sheets.properties")
        .execute()
    )
    titulos = [s["properties"]["title"] for s in meta.get("sheets", [])]
    return _escolher_aba_com_cabecalho(titulos, carregar)


def _ler_linhas_excel(
    drive, spreadsheet_id: str, aba: str | None
) -> tuple[str, list[list[str]]]:
    conteudo = (
        drive.files()
        .get_media(fileId=spreadsheet_id, supportsAllDrives=True)
        .execute()
    )
    workbook = openpyxl.load_workbook(
        io.BytesIO(conteudo), read_only=True, data_only=True
    )
    try:
        def carregar(titulo: str) -> list[list[str]]:
            planilha = workbook[titulo]
            linhas: list[list[str]] = []
            for indice, linha in enumerate(planilha.iter_rows(values_only=True)):
                if indice >= _MAX_LINHAS:
                    break
                linhas.append([_celula_excel(celula) for celula in linha])
            return linhas

        if aba is not None:
            return aba, carregar(aba)
        return _escolher_aba_com_cabecalho(workbook.sheetnames, carregar)
    finally:
        workbook.close()


def _escolher_aba_com_cabecalho(
    titulos: list[str],
    carregar_linhas,
) -> tuple[str, list[list[str]]]:
    """Escolhe a aba certa pelo conteudo, nao pelo nome.

    A aba do Projeto Funcional e identificada por ter uma linha de cabecalho
    com 'Status' e 'Descricao' (o nome da aba costuma ser o do cliente).
    """
    if not titulos:
        raise PlanilhaProjetoInvalida("A planilha nao possui nenhuma aba.")

    if len(titulos) == 1:
        return titulos[0], carregar_linhas(titulos[0])

    for titulo in titulos:
        linhas = carregar_linhas(titulo)
        if _localizar_linha_cabecalho(linhas) is not None:
            return titulo, linhas

    # Nenhuma aba com cabecalho reconhecido: usa a primeira como fallback.
    return titulos[0], carregar_linhas(titulos[0])


def _celula_excel(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, datetime.datetime):
        return valor.strftime("%d/%m/%Y")
    if isinstance(valor, datetime.date):
        return valor.strftime("%d/%m/%Y")
    return str(valor).strip()


# ---------------------------------------------------------------------------
# Parsing comum (cabecalho + linhas)
# ---------------------------------------------------------------------------


def _normalizar(texto: str) -> str:
    texto = (texto or "").strip().lower()
    decomposto = unicodedata.normalize("NFD", texto)
    return "".join(c for c in decomposto if unicodedata.category(c) != "Mn")


def _campo_do_cabecalho(celula: str) -> str | None:
    """Mapeia um texto de cabecalho para o nome interno do campo."""
    n = _normalizar(celula)
    if not n:
        return None
    if n in {"#", "n", "no", "nº", "item", "id"}:
        return "numero"
    if "registro" in n or n == "data":
        return "data_registro"
    if "modulo" in n:
        return "modulo"
    if "funcao" in n or "processo" in n:
        return "funcao"
    if "analista" in n:
        return "analista"
    if "descricao" in n:
        return "descricao"
    if "prioridade" in n:
        return "prioridade"
    if n == "status" or n.startswith("status"):
        return "status"
    return None


def _localizar_linha_cabecalho(linhas: list[list[str]]) -> int | None:
    for indice, linha in enumerate(linhas):
        campos = {_campo_do_cabecalho(celula) for celula in linha}
        if "status" in campos and "descricao" in campos:
            return indice
    return None


def _mapear_colunas(linha_cabecalho: list[str]) -> dict[str, int]:
    mapa: dict[str, int] = {}
    for indice, celula in enumerate(linha_cabecalho):
        campo = _campo_do_cabecalho(celula)
        if campo and campo not in mapa:
            mapa[campo] = indice
    return mapa


def _linha_para_item(
    linha: list[str],
    mapa: dict[str, int],
    sequencial: int,
) -> ItemProjetoFuncional | None:
    status = _valor(linha, mapa.get("status"))
    descricao = _valor(linha, mapa.get("descricao"))
    if not status and not descricao:
        return None

    numero = _valor(linha, mapa.get("numero")) or str(sequencial)
    return ItemProjetoFuncional(
        numero=numero,
        data_registro=_valor(linha, mapa.get("data_registro")),
        modulo=_valor(linha, mapa.get("modulo")),
        funcao=_valor(linha, mapa.get("funcao")),
        analista=_valor(linha, mapa.get("analista")),
        descricao=descricao,
        prioridade=_valor(linha, mapa.get("prioridade")),
        status=status,
    )


def _valor(linha: list[str], indice: int | None) -> str:
    if indice is None or indice >= len(linha) or linha[indice] is None:
        return ""
    return str(linha[indice]).strip()
